import utils
import logging

from merge_data import load_fips, state_fips_to_name
from openpyxl import load_workbook

log = logging.getLogger(__name__)
log.addHandler(logging.StreamHandler())

FIPS_TO_STATE = {}
STATE_TO_FIPS = {}

def load_hud_crosswalk(fn):
    book = load_workbook(fn)
    sheet = book.active

    zccd = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        z = str(row[0]).replace('="', '').replace('"', '')
        stcd = str(row[1]).replace('="', '').replace('"', '')
        st = stcd[:2]
        cd = stcd[2:]

        if cd == '**':
            msg = 'invalid CD for %s: %s' % (z, cd)
            log.error(msg)
            continue

        try:
            zccd.append({
                'zip': z,
                'state_fips': st,
                'state_abbr': FIPS_TO_STATE[st],
                'cd': str(int(cd)) # string conversion to drop leading zero
            })
        except Exception as e:
            msg = 'unable to convert CD for %s: %s, error: %s' % (z, stcd, str(e))
            log.error(msg)
            continue
    return zccd


if __name__ == "__main__":
    # load state FIPS codes
    FIPS_TO_STATE = load_fips('raw/state_fips.txt')
    STATE_TO_FIPS = {v: k for k, v in FIPS_TO_STATE.items()}

    # load HUD crosswalk file
    zccd_hud = load_hud_crosswalk('raw/hud_crosswalk.xlsx')

    # sort by fips
    zccd_sorted = sorted(zccd_hud, key=lambda k: (k['state_fips'], k['zip'], k['cd']))

    # write output
    utils.csv_writer('zccd_hud.csv', zccd_sorted, ['state_fips', 'state_abbr', 'zip', 'cd'])